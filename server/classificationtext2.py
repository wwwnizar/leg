import json as j
import pandas as pd
import re
import numpy as np
from nltk.corpus import stopwords
from nltk.stem.snowball import FrenchStemmer
from nltk.stem import SnowballStemmer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import LinearSVC
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import SelectKBest, chi2
from sklearn.feature_selection import SelectPercentile
import openpyxl
from openpyxl import *
from openpyxl import load_workbook
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer
from nltk.corpus import stopwords
from nltk.stem.snowball import FrenchStemmer
from cltk.corpus.utils.formatter import normalize_fr
from cltk.stop.french.stops import STOPS_LIST as FRENCH_STOPS
from cltk.tokenize.word import WordTokenizer
from cltk.corpus.utils.formatter import normalize_fr
from cltk.lemmatize.french.lemma import LemmaReplacer
from sklearn import tree
from sklearn.tree                   import DecisionTreeClassifier
import graphviz 
import pydotplus
import collections


json_data = None
tokenizer = WordTokenizer('french')
wb = load_workbook(filename='classif_questions.xlsx')

# On active l'onglet courant
ws = wb['Feuil2']

# On crГ©e un nouvel onglet
ws1 = wb.create_sheet()
ws1.title = ws.title
stemmer = FrenchStemmer()
words = set(stopwords.words("french"))  # load stopwords
words.add("la")
words.add("le")
words.add("dans")
words.add("un")
words.add("une")
words.add("a")
lstwords=list(words)
# ouverture du fichier Excel 
joined_lines=[]
lines=[]
for c in range(2,200):
    text = ws.cell(row=c,column=2).value
    target= str(ws.cell(row=c,column=3).value)
    joined_lines.append([text ,target])




data = pd.DataFrame(joined_lines,columns=['text', 'stars'])
#print (data)



data['cleaned']=data['text'].apply(lambda x: " ".join([stemmer.stem(i) for i in re.sub("[^a-zA-Z]", " ",x).split() if i not in words]).lower())
#print (data['cleaned'])
X_train, X_test, y_train,y_test = train_test_split(data['cleaned'], data.stars, test_size=0.2)
pipeline= Pipeline([('vect', TfidfVectorizer(ngram_range=(1, 2), stop_words= lstwords, sublinear_tf=True)),
                    ('chi', SelectKBest(chi2, k = 3000)),#SelectPercentile(percentile=90)),#
                    ('clf', tree.DecisionTreeClassifier())])

ytrain = []
for i in y_train:
    ytrain.append(str(i))
model = pipeline.fit(X_train,ytrain)
vectorizer= model.named_steps['vect']
chi = model.named_steps['chi']
clf = model.named_steps['clf']
feature_names = vectorizer.get_feature_names()
feature_names=[feature_names[i] for i in chi.get_support(indices = True)]
feature_names = np.asarray(feature_names)




dot_data = tree.export_graphviz(clf,
                                feature_names=feature_names,
                                out_file=None,
                                filled=True,
                                rounded=True)
graph = pydotplus.graph_from_dot_data(dot_data)

colors = ('turquoise', 'orange')
edges = collections.defaultdict(list)

for edge in graph.get_edge_list():
    edges[edge.get_source()].append(int(edge.get_destination()))

for edge in edges:
    edges[edge].sort()    
    for i in range(2):
        dest = graph.get_node(str(edges[edge][i]))[0]
        dest.set_fillcolor(colors[i])

graph.write_png('tree.png')


ytest = []
for i in y_test:
    ytest.append(str(i))
print("accuracy score : "+str(model.score(X_test, ytest)))

text1= "Bonjour, Me confirmez-vous que la formation en PJ est bien remboursГ©e en frais de mission et non en frais de stage? Je vous remercie."
#0
text2= "l'adj lavenu devant effectuer le Stage de formation des inspecteurs de sГ©curitГ© de la dГ©fense Г  malakoff du 05/01/15 au 17/04/2015 nous avons rГ©alisГ© un OM avec avance avec comme condition un hГ©bergement dans le secteur privГ© puisque le centre de formation ne dispose pas d'hГ©bergement. Or le CAMID a traitГ© le dossier en IJS contrairement Г  notre demande. Est-il possible de rГ©viser son dossier afin de lui faire un complГ©ment d'avance puisque celle-ci est moindre par rapport a son stage de 4 mois en rГ©gion parisienne? ci-joint un exemplaire de son attestation de stage."
#3
text3= "L'ADMINISTRE A EFFECTUE UN STAGE A TOULON DU 17/09/2014 AU 22/11/2014. PENDANT CE STAGE IL A EFFECTUE PLUSIEURS ALLER/RETOUR ENTRE SON LIEU DE STAGE ET SON DOMICILE. LES JUSTIFICATIFS AFFERENTS ONT ETE FOURNIS. DORIA NE CALCULE PAS LES DIFFERENTS SEJOURS EN FONCTION DE CES TRAJETS, A LA CREATION DU DORIA OU EN LIQUIDATION. DE CE FAIT PEUT-ON SCANNER UN ETAT RECAPITULATIF DE CES TRAJETS EN PIECE JOINTE SANS CREER LES TRAJETS SUR LE DORIA ?"
#1
print (model.predict([text1]))
print (model.predict([text2]))
print (model.predict([text3]))

#######################################################################################################

text = "Bonjour, Un stage qui a été effectué du 19/10 au 12/12/2014 et qui comporte des stages son considérée comme mission donc pas d'attestation de stage à fournir ? Pourriez-vous m'indiquer la marche à suivre pour ce genre de cas ? dois-je les codifier seulement en stage (taux mission) ou en mission ?"
text = (" ".join([stemmer.stem(i) for i in re.sub("[^a-zA-Z]", " ",text).split() if i not in words]).lower())
tokenizer = WordTokenizer('french')
token_item = tokenizer.tokenize(text) 
print (token_item)
tf = TfidfVectorizer(ngram_range=(1, 2), stop_words= lstwords, sublinear_tf=True)
X = tf.fit_transform(token_item)
idf = tf.idf_
#print(tf.vocabulary_)
output = {}
output["Output"]=[]
output["Output"].append({
    'Vecteur': ",".join(token_item),
    'classe': str(model.predict([text]))
    })


#print (output)
with open('data.txt', 'w') as outfile:  
    j.dump(output, outfile)